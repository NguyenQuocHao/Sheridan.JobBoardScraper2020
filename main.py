# Original work Copyright (c) 2020 [Hao Nguyen]

import openpyxl
from bs4 import BeautifulSoup
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time

# Getting things ready!!

# Check if excel file exists, if not create one
file_name = 'sheridan_job_board_fall_2020_v3.xlsx'
FileExistence = Path(file_name)
if FileExistence.is_file():
    wb = openpyxl.load_workbook(file_name)
else:
    wb = openpyxl.Workbook()  # Create a blank workbook.
sheet = wb.active

# from webdriver_manager.chrome import ChromeDriverManager
# driver = webdriver.Chrome(ChromeDriverManager().install())
browser = webdriver.Chrome()
browser.get('https://jobs.sheridancollege.ca/Student/Default.asp#Content')
time.sleep(1)  # Sleep to wait for the browser to load the content

# Fill in login info.
# Account
userElem = browser.find_element_by_id('user')
userElem.send_keys("")  # Put your studentID in ""
# Password
passwordElem = browser.find_element_by_id('password')
passwordElem.send_keys("")  # Put your password in ""
htmlElem = browser.find_element_by_tag_name('html')
passwordElem.send_keys(Keys.ENTER)

# Navigate to Job board tab
linkElem = browser.find_element_by_link_text('Job Postings')
linkElem.click()

# Set headers for table
header_list = [
    "Job Number",
    "Employer",
    "Job Title",
    "Status",
    "Job Type",
    "Job Sub-Type",
    "City",
    "Positions",
    "Months",
    "Posting Date",
    "Closing Date",
    "Salary",
    "Compensation",
    "Description"
]

last_page_exception_message = "Reached last page."
count = 0  # Number of moves (cell to cell)
countX = 1  # Column
countY = 2  # Row
job_id_top = ''


def create_headers(sheet, headers):
    column = 1
    for header in headers:
        sheet.cell(row=1, column=column).value = header
        column += 1


# Get job board table content
def collect_job_postings():
    soup = BeautifulSoup(browser.page_source, 'lxml')  # load page content
    my_table = soup.find('table', {'class': 'table table-bordered'})
    found_tds = my_table.find_all('td')  # scan through table, and store all 'td' elements into an array
    global job_id_top
    for found_td in found_tds:
        global count
        global countX
        global countY
        global last_page_exception_message
        # copy salary & compensation
        if count % 11 == 0:  # execute if first cell
            job_id = found_td.text.replace('\n', '').strip()
            locator = browser.find_element_by_link_text(job_id)

            # if next page is the previous page -> escape and throw exception
            if job_id_top == job_id:
                raise Exception(last_page_exception_message)

            # assign job id at the top
            if (countY - 2) % 20 == 0:
                job_id_top = job_id

            sheet.cell(row=countY, column=countX).value = job_id
            locator.click()
            soup = BeautifulSoup(browser.page_source, 'lxml')  # update buffered link to current link
            my_salary = soup.find('div', {'id': 'jd9'})
            my_compensation = soup.find('div', {'id': 'jd10'})
            my_description = soup.find('div', {'id': 'jobd1'})
            if my_salary is not None:
                find_salary = my_salary.find_all('p')[0].text.replace('\n', '').strip()
                sheet.cell(row=countY, column=12).value = find_salary
            if my_compensation is not None:
                find_compensation = my_compensation.find_all('p')[0].text.replace('\n', '').strip()
                sheet.cell(row=countY, column=13).value = find_compensation
            if my_description is not None:
                find_description = my_description.find_all('p')[0].text.replace('\n', '').strip()
                sheet.cell(row=countY, column=14).value = find_description
            browser.back()

        elif count % 11 != 0:  # execute if not first cell
            sheet.cell(row=countY, column=countX).value = found_td.text.replace('\n', '').strip()
            # Check if the last cell
            if count % 11 == 10:
                countY += 1
                countX = 0

        # Increment anyways
        countX += 1
        count += 1


# Create headers for the table
create_headers(sheet, header_list)

# loop through pages
for page in range(0, 100):
    try:
        collect_job_postings()
    except Exception as error:
        if str(error) == last_page_exception_message:
            break
    # Go to the next page
    linkElem = browser.find_element_by_link_text('Next')
    linkElem.click()


# Write (save) the excel file into the drive
wb.save(file_name)
